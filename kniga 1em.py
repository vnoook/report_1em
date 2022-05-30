# Программа для создания отчёта 1-ЕМ.
#
# В папке с программой должен быть пустой шаблон 1-ЕМ и папка 'res',
# а в ней должны находиться файлы отчётов из АРМ Статистика.
#
# Программа пропускает 4 строки из переменной 'rows_for_skip'.
# После отработки программы остаётся поменять на титульном листе период.
#
#     Максим Красовский \ февраль 2021 (май 2022) \ noook@yandex.ru

import time
import datetime
import openpyxl

# считаю время скрипта
time_start = time.time()
print('начинается' + '.'*20)


# функция для анализа что выдавать в ячейку
# в ячейке может быть целое или дробное число, строка, пусто
def conv_cell(cell_value):
    if type(cell_value) == int:
        return cell_value
    else:
        if cell_value == '***' or cell_value == '0,0':
            return cell_value
        elif cell_value is None:
            cell_value = ''
            return cell_value
        else:
            cell_value = float(cell_value.replace(',', '.'))
            return cell_value


# функция для анализа что в ячейке
# если в ячейке то, что можно преобразовать в число, то выдать иначе выдать False
def int_cell(cell_value):
    if type(cell_value) == int:
        return cell_value
    else:
        return False


# функция составления названия файла для сохранения
# состоит из названия отчёта + месяца + года
def name_of_file():
    # текущие месяц и год
    number_of_month = datetime.datetime.today().month
    number_of_year = datetime.datetime.today().year

    # если запустили в январе, то (месяц и год) надо сменить на (декабрь и (год-1))
    # иначе (месяц-1)
    if number_of_month == 1:
        number_of_month = 12
        number_of_year -= 1
    else:
        number_of_month -= 1

    # если номер месяца цифра, то добавить 0 в начало
    # иначе просто перевести в строку
    if number_of_month < 10:
        name_month = '0'+str(number_of_month)
    else:
        name_month = str(number_of_month)

    file_name = '1-em-' + name_month + '-' + str(number_of_year) + '.xlsx'
    return file_name


# функция составления строки периода отчёта
# состоит из 'январь' + текущий месяц
def name_of_period():
    # 'декабрь' задублирован для того, что бы не делать (месяц-1) дальше по алгоритму
    month_tuple = ('декабрь', 'январь', 'февраль', 'март', 'апрель', 'май', 'июнь',
                   'июль', 'август', 'сентябрь', 'октябрь', 'ноябрь', 'декабрь'
                   )
    # текущие месяц и год
    number_of_month = datetime.datetime.today().month
    number_of_year = datetime.datetime.today().year

    # если запустили в январе, то (месяц и год) надо сменить на (декабрь и (год-1))
    if number_of_month == 1:
        name_month = month_tuple[0]
        name_year = str(number_of_year - 1)
    else:
        name_month = month_tuple[number_of_month - 1]
        name_year = str(number_of_year)

    str_period = month_tuple[1] + ' - ' + name_month + ' ' + name_year
    return str_period


# файлы для работы
xl_template = 'ШАБЛОН 1-em-05-2022.xlsx'
xl_1em_sheets = {
                '1': 'res/01-1 ст.xlsx',
                '2': 'res/02Принято к производству дел в отчетном периоде.xlsx',
                '3': 'res/03Отменено постановлений о возбуждении уг. дела .xlsx',
                '4': 'res/04Всего окончено дел в отчетном периоде (с повторными).xlsx',
                '5': 'res/05Направлено дел прокурору ... мер медицинского характера.xlsx',
                '6': 'res/06с обвинительным заключением либо актом.xlsx',
                '7': 'res/07возвращено прокурором дел для производства доп.расследования.xlsx',
                '8': 'res/08Число обвиняемых по направленным в суд делам.xlsx',
                '9': 'res/09Поступило от прокурора дел, возвращенных судом в порядке ст.237 УПК РФ.xlsx',
                '10': 'res/10Прекращено дел (с повторными).xlsx',
                '11': 'res/11Число лиц, в отношении которых прекращены дела иили уголовное преследование.xlsx',
                '12': 'res/12Приостановлено дел производством в отчетном периоде.xlsx',
                '13': 'res/13ввиду неустанов. лица, подлежащего привлеч. в качестве обвиняемого.xlsx',
                '14': 'res/14ввиду неустановления места нахождения подозреваемого или обвиняемого.xlsx',
                '15': 'res/15ввиду отсутствия реальной ... место нахождения которого известно.xlsx',
                '16': 'res/16ввиду временного тяжелого заболевания подозреваемого или обвиняемого.xlsx',
                '17': 'res/17Расследовано дел в сроки свыше установленного УПК РФ.xlsx',
                '18': 'res/18Остаток неоконченных дел на конец месяца.xlsx',
                '19': 'res/19Число лиц, в отношении которых ... прекращено за  непричастностью.xlsx',
                '20': 'res/20Число оправданных и лиц, дела о  которых прекращены судом ... в связи с непричастностью.xlsx'
                }

# строки для пропуска
rows_for_skip = (52, 53, 54, 60)

# переменные для работы
start_row_first_page = 4
start_col_first_page = 3
max_row_first_page = 24
max_col_first_page = 18
max_row_another_page = 61
max_col_another_page = 25

# открываю книгу шаблон в которую вставляю данные
wb_1em = openpyxl.load_workbook(xl_template)

# иду по листам шаблона чтобы вставить данные из файлов
# беру все листы шаблона по очереди
# wb_1em       - файл шаблона, wb_1em_s       - лист в шаблоне
# wb_file_data - файл шаблона, wb_file_data_s - лист в шаблоне
for dict_key in xl_1em_sheets:
    # назначаю в шаблоне активный лист
    wb_1em_s = wb_1em[dict_key]

    # открываю книгу из которой беру данные
    wb_file_data = openpyxl.load_workbook(xl_1em_sheets[dict_key])
    wb_file_data_s = wb_file_data.active

    print('\n' + xl_1em_sheets[dict_key])

    # постраничный алгоритм, на каждом листе по своему считается
    # первый лист особый подход
    if wb_1em.index(wb_1em_s) == 1:
        for i_row in range(start_row_first_page, max_row_first_page+1):
            for i_col in range(start_col_first_page, max_col_first_page+1):
                # C4:R24 -> D5:S25  ||  R4C3:R24C18 -> R5C4:R25C19
                wb_1em_s.cell(i_row+1, i_col+1).value = conv_cell(wb_file_data_s.cell(i_row, i_col).value)

    # остальные листы все по одному алгоритму
    # пропускаются строки с номерами из rows_for_skip
    else:
        # алгоритм обновления ячеек с периодом отчёта
        # A5 || R5:C1
        wb_1em_s.cell(5, 1).value = name_of_period()

        # переменная сдвига для поднятия данных после пропуска строки
        i_shift = 0

        for i_row in range(8, max_row_another_page+1):
            if i_row in rows_for_skip:
                i_shift += 1
            else:
                for i_col in range(2, max_col_another_page+1):
                    # B8:Y61 -> B8:Y57 || R8C2:R61C25 -> R8C2:R57C25
                    wb_1em_s.cell(i_row-i_shift, i_col).value = conv_cell(wb_file_data_s.cell(i_row, i_col).value)

        # алгоритм пересчёта некоторых ячеек чисто для отчёта 1-ЕМ
        # тут обрабатываются трёхколоночные листы
        if int(dict_key) in (2, 3, 4, 5, 8, 11, 12, 18, 19, 20):
            for i_row in range(rows_for_skip[1], rows_for_skip[2]+1):
                for i_col in range(2, max_col_another_page+1):
                    if int_cell(wb_file_data_s.cell(i_row, i_col).value):
                        if i_col % 3 in (2, 0):
                            wb_1em_s.cell(i_row-40, i_col).value = wb_file_data_s.cell(i_row, i_col).value +\
                                                                   wb_file_data_s.cell(i_row-40, i_col).value
                            if i_col % 3 == 0:
                                decision_cell =\
                                    wb_1em_s.cell(i_row-40, i_col-1).value - wb_1em_s.cell(i_row-40, i_col).value
                                wb_1em_s.cell(i_row-40, i_col+1).value =\
                                    (decision_cell/wb_1em_s.cell(i_row-40, i_col).value)*100

        # тут обрабатываются четырёхколоночные листы
        elif int(dict_key) in (6, 7, 9, 10, 13, 14, 15, 16, 17):
            # побежал по ячейкам
            for i_row in range(rows_for_skip[1], rows_for_skip[2]+1):
                for i_col in range(2, max_col_another_page+1):
                    # если содержимое число, то продолжаю
                    if int_cell(wb_file_data_s.cell(i_row, i_col).value):
                        # если это нужные колонки, то обрабатываю
                        if i_col % 4 in (2, 3):
                            wb_1em_s.cell(i_row-40, i_col).value = wb_file_data_s.cell(i_row, i_col).value +\
                                                                   wb_file_data_s.cell(i_row-40, i_col).value

    # закрываю файл из которого беру данные
    wb_file_data.close()

# сохраняю файл шаблона и закрываю его
wb_1em.save(name_of_file())
wb_1em.close()

# считаю время скрипта
time_finish = time.time()
print('\n' + '.'*30 + 'закончено за', round(time_finish-time_start, 3), 'секунд')

# закрываю программу
input('\nНажмите ENTER')
